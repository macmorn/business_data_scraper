#!/usr/bin/env python3
"""Export enriched company CSV to a formatted Excel workbook.

Produces two sheets:
1. "Company Overview" — key business information
2. "Detailed Data" — auxiliary financials, metadata, and JSON fields

Usage:
    python export_excel.py [--input path/to/input.csv] [--output path/to/output.xlsx]
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import OrderedDict
from pathlib import Path

import pandas as pd
import pgeocode
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single")

MIN_COL_WIDTH = 8
MAX_COL_WIDTH = 50
LONG_TEXT_MAX = 40

# ---------------------------------------------------------------------------
# Column mappings  (csv_column -> Excel header)
# ---------------------------------------------------------------------------
SHEET1_COLUMNS: OrderedDict[str, str] = OrderedDict([
    ("company_name_original", "Company Name"),
    ("matched_name", "Matched Name"),
    ("country", "Country"),
    ("legal_form", "Legal Form"),
    ("status", "Status"),
    ("address", "Address"),
    ("region", "Bundesland / Region"),
    ("register_id", "Register ID"),
    ("register_court", "Register Court"),
    ("lei", "LEI"),
    ("vat_id", "VAT ID"),
    ("industry_code", "Industry Code"),
    ("corporate_purpose", "Corporate Purpose"),
    ("founded_year", "Founded"),
    ("employees_count", "Employees"),
    ("employees_range", "Employee Range"),
    ("revenue", "Revenue"),
    ("revenue_range", "Revenue Range"),
    ("earnings", "Earnings"),
    ("total_assets", "Total Assets"),
    ("equity", "Equity"),
    ("ceo_name", "CEO Name"),
    ("ceo_current_title", "CEO Title"),
    ("ceo_linkedin_url", "CEO LinkedIn"),
    ("ceo_career_summary", "CEO Career Summary"),
    ("ceo_confidence", "CEO Confidence"),
    ("confidence_score", "Confidence"),
    ("needs_review_flag", "Needs Review"),
])

SHEET2_COLUMNS: OrderedDict[str, str] = OrderedDict([
    ("company_name_original", "Company Name"),
    ("employees_notes", "Employee Notes"),
    ("revenue_notes", "Revenue Notes"),
    ("equity_ratio", "Equity Ratio (%)"),
    ("return_on_sales", "Return on Sales (%)"),
    ("cost_of_materials", "Cost of Materials"),
    ("wages_and_salaries", "Wages & Salaries"),
    ("cash_on_hand", "Cash on Hand"),
    ("liabilities", "Liabilities"),
    ("pension_provisions", "Pension Provisions"),
    ("auditor", "Auditor"),
    ("public_funding_total", "Public Funding Total"),
    ("last_accounts_year", "Last Accounts Year"),
    ("officers", "Officers (JSON)"),
    ("corporate_structure_summary", "Corporate Structure"),
    ("northdata_url", "Northdata URL"),
    ("data_sources_used", "Data Sources"),
    ("financials_json", "Financials (JSON)"),
])

# Column categories for formatting
FINANCIAL_COLS = {
    "total_assets", "equity", "cost_of_materials", "wages_and_salaries",
    "cash_on_hand", "liabilities", "pension_provisions", "public_funding_total",
}
MIXED_FINANCIAL_COLS = {"revenue", "earnings"}
YEAR_COLS = {"founded_year", "last_accounts_year"}
URL_COLS = {"ceo_linkedin_url", "northdata_url"}
LONG_TEXT_COLS = {
    "corporate_purpose", "ceo_career_summary", "officers",
    "corporate_structure_summary", "financials_json", "address",
}


# ---------------------------------------------------------------------------
# Region lookup via postal code
# ---------------------------------------------------------------------------
_PLZ_RE = re.compile(r"(?<!\d)\b(\d{4,5})\b")
_NOMINATIM_CACHE: dict[str, pgeocode.Nominatim] = {}

# Country-specific region label for the Excel header
REGION_LABELS = {
    "DE": "Bundesland",
    "AT": "Bundesland",
    "CH": "Kanton",
    "NL": "Provincie",
    "BE": "Province / Région",
    "FR": "Région",
    "ES": "Comunidad Autónoma",
    "LU": "Luxembourg",
}


def _get_nominatim(country_code: str) -> pgeocode.Nominatim | None:
    """Return a cached pgeocode.Nominatim instance for *country_code*."""
    cc = country_code.upper()
    if cc not in _NOMINATIM_CACHE:
        try:
            _NOMINATIM_CACHE[cc] = pgeocode.Nominatim(cc)
        except Exception:
            _NOMINATIM_CACHE[cc] = None  # type: ignore[assignment]
    return _NOMINATIM_CACHE[cc]


def _resolve_region(address: str | None, country: str | None) -> str | None:
    """Extract postal code from *address* and return the state/region name."""
    if not address or not country or pd.isna(address) or pd.isna(country):
        return None
    address, country = str(address), str(country)
    cc = country.strip().upper()[:2]
    nomi = _get_nominatim(cc)
    if nomi is None:
        return None
    m = _PLZ_RE.search(address)
    if not m:
        return None
    result = nomi.query_postal_code(m.group(1))
    state = getattr(result, "state_name", None)
    if state is None or (isinstance(state, float) and pd.isna(state)):
        return None
    return str(state)


def _add_region_column(df: pd.DataFrame) -> pd.DataFrame:
    """Add a 'region' column derived from address + country."""
    df["region"] = df.apply(
        lambda r: _resolve_region(r.get("address"), r.get("country")), axis=1
    )
    return df


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _try_numeric(value):
    """Try to convert a value to float; return original if it fails."""
    if pd.isna(value):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return str(value)


def _write_sheet(ws, df: pd.DataFrame, col_map: OrderedDict[str, str]) -> None:
    """Write data to a worksheet with proper types."""
    csv_cols = list(col_map.keys())
    headers = list(col_map.values())

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Write data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, csv_col in enumerate(csv_cols, start=1):
            raw = row.get(csv_col)
            if pd.isna(raw):
                continue

            cell = ws.cell(row=row_idx, column=col_idx)

            if csv_col in FINANCIAL_COLS or csv_col in MIXED_FINANCIAL_COLS:
                val = _try_numeric(raw)
                cell.value = val
                if isinstance(val, float):
                    cell.number_format = '#,##0'
            elif csv_col == "employees_count":
                val = _try_numeric(raw)
                cell.value = val
                if isinstance(val, float):
                    cell.number_format = '#,##0'
            elif csv_col == "confidence_score":
                val = _try_numeric(raw)
                cell.value = val
                if isinstance(val, float):
                    cell.number_format = '0%'
            elif csv_col in {"equity_ratio", "return_on_sales"}:
                val = _try_numeric(raw)
                cell.value = val
                if isinstance(val, float):
                    cell.number_format = '0.0"%"'
            elif csv_col in YEAR_COLS:
                val = _try_numeric(raw)
                if isinstance(val, float):
                    cell.value = int(val)
                    cell.number_format = '0'
                else:
                    cell.value = val
            elif csv_col in URL_COLS:
                url = str(raw)
                cell.value = url
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.font = LINK_FONT
            elif csv_col == "needs_review_flag":
                cell.value = str(raw).strip().lower() in ("true", "1", "yes")
            else:
                cell.value = str(raw)

    # Freeze header row
    ws.freeze_panes = "A2"


def _auto_fit_widths(ws, col_map: OrderedDict[str, str]) -> None:
    """Auto-fit column widths with min/max constraints."""
    csv_cols = list(col_map.keys())
    for col_idx in range(1, len(csv_cols) + 1):
        csv_col = csv_cols[col_idx - 1]
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))

        # Sample up to 200 rows to estimate width
        for row_idx in range(2, min(ws.max_row + 1, 202)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, min(len(str(val)), MAX_COL_WIDTH))

        cap = LONG_TEXT_MAX if csv_col in LONG_TEXT_COLS else MAX_COL_WIDTH
        width = max(MIN_COL_WIDTH, min(max_len + 2, cap))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Enable wrap for long-text columns
        if csv_col in LONG_TEXT_COLS:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _add_review_highlighting(ws, col_map: OrderedDict[str, str]) -> None:
    """Highlight rows where needs_review_flag is True."""
    csv_cols = list(col_map.keys())
    if "needs_review_flag" not in csv_cols:
        return

    review_col_idx = csv_cols.index("needs_review_flag") + 1
    review_letter = get_column_letter(review_col_idx)
    last_col_letter = get_column_letter(len(csv_cols))
    last_row = ws.max_row
    if last_row < 2:
        return

    formula = f"${review_letter}2=TRUE"
    ws.conditional_formatting.add(
        f"A2:{last_col_letter}{last_row}",
        FormulaRule(formula=[formula], fill=REVIEW_FILL),
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export enriched CSV to formatted Excel workbook.",
    )
    parser.add_argument(
        "--input", "-i",
        default="output/companies_enriched.csv",
        help="Path to input CSV (default: output/companies_enriched.csv)",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Path to output XLSX (default: input path with .xlsx extension)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else input_path.with_suffix(".xlsx")

    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    # Read CSV
    df = pd.read_csv(input_path, encoding="utf-8-sig", dtype=str, keep_default_na=True)
    print(f"Read {len(df)} companies from {input_path}")

    # Derive region from postal code in address
    _add_region_column(df)
    region_count = df["region"].notna().sum()
    print(f"Resolved region for {region_count}/{len(df)} companies")

    # Create workbook
    wb = Workbook()

    # Sheet 1: Company Overview
    ws1 = wb.active
    ws1.title = "Company Overview"
    _write_sheet(ws1, df, SHEET1_COLUMNS)
    _auto_fit_widths(ws1, SHEET1_COLUMNS)
    _add_review_highlighting(ws1, SHEET1_COLUMNS)

    # Sheet 2: Detailed Data
    ws2 = wb.create_sheet("Detailed Data")
    _write_sheet(ws2, df, SHEET2_COLUMNS)
    _auto_fit_widths(ws2, SHEET2_COLUMNS)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Exported to {output_path}")


if __name__ == "__main__":
    main()
